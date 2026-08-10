import io
import logging
from typing import Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelWorkbookExporter:
    """
    Multi-Tab Excel Workbook Exporter using OpenPyXL.
    Produces formatted enterprise spreadsheets with executive summary tabs.
    """

    def generate_excel_workbook(
        self,
        filename: str,
        df: pd.DataFrame,
        kpis: Dict[str, Any],
        summary_stats: Dict[str, Any] = None
    ) -> bytes:
        wb = openpyxl.Workbook()
        
        # Header Style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # ── Sheet 1: Executive Summary ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True

        ws1.cell(row=1, column=1, value="AI Data Dashboard — Executive Summary").font = title_font
        ws1.cell(row=2, column=1, value=f"Dataset Source: {filename}").font = subtitle_font

        ws1.cell(row=4, column=1, value="Key Performance Metric").font = header_font
        ws1.cell(row=4, column=1).fill = header_fill
        ws1.cell(row=4, column=2, value="Value").font = header_font
        ws1.cell(row=4, column=2).fill = header_fill

        kpi_rows = [
            ("Total Orders Processed", kpis.get("total_orders", 0)),
            ("Total Revenue", f"£{kpis.get('total_revenue', 0.0):,.2f}"),
            ("Total Profit", f"£{kpis.get('total_profit', 0.0):,.2f}"),
            ("Average Order Value (AOV)", f"£{kpis.get('average_order_value', 0.0):,.2f}"),
        ]

        for idx, (label, val) in enumerate(kpi_rows, start=5):
            c1 = ws1.cell(row=idx, column=1, value=label)
            c2 = ws1.cell(row=idx, column=2, value=val)
            c1.font = bold_font
            c1.border = thin_border
            c2.border = thin_border

        # ── Sheet 2: Cleaned Dataset ───────────────────────────────────────────
        ws2 = wb.create_sheet(title="Cleaned Dataset")
        ws2.views.sheetView[0].showGridLines = True

        # Write Headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write Data Rows
        for row_idx, row in enumerate(df.head(5000).itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                c = ws2.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                c.border = thin_border

        # Auto-adjust Column Widths across all sheets
        for sheet in [ws1, ws2]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes
